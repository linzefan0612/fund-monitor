# -*- coding: utf-8 -*-
"""
基金实时监控助手 - Flask Web应用
提供Web界面和API接口
"""

import os
import json
import logging
from datetime import datetime
from flask import Flask, render_template, jsonify, request

from config import REFRESH_INTERVAL, AFTERNOON_REMINDER_TIME
from monitor import monitor, Alert
from advisor import init_advisor, get_advisor
from fund_data import data_source_manager

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# 创建Flask应用
app = Flask(__name__)
app.config['JSON_AS_ASCII'] = False  # 支持中文

# 初始化建议模块
init_advisor(monitor)


# ==================== 页面路由 ====================

@app.route('/')
def index():
    """主页 - Vue版本"""
    return render_template('index_vue.html', refresh_interval=REFRESH_INTERVAL)


@app.route('/element')
def index_element():
    """主页 - Element Plus版本"""
    return render_template('index_element.html', refresh_interval=REFRESH_INTERVAL)


# ==================== API路由 ====================

@app.route('/api/funds', methods=['GET'])
def get_funds():
    """获取所有监控的基金数据"""
    try:
        funds_data = monitor.get_all_funds_data()
        result = []
        for code, data in funds_data.items():
            position = monitor.get_position(code)
            risk = monitor.get_risk_level(code)

            result.append({
                **data,
                'current_amount': position.current_amount if position else 0,
                'max_amount': position.max_amount if position else 0,
                'holding_return_rate': position.holding_return_rate if position else 0,
                'risk_level': risk
            })

        return jsonify({
            'success': True,
            'data': result,
            'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
    except Exception as e:
        logger.error(f"获取基金数据失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/funds/refresh', methods=['POST'])
def refresh_funds():
    """刷新所有基金数据"""
    try:
        data = monitor.refresh_data()
        result = []
        for code, fund_data in data.items():
            position = monitor.get_position(code)
            risk = monitor.get_risk_level(code)

            result.append({
                **fund_data,
                'current_amount': position.current_amount if position else 0,
                'max_amount': position.max_amount if position else 0,
                'holding_return_rate': position.holding_return_rate if position else 0,
                'risk_level': risk
            })

        return jsonify({
            'success': True,
            'data': result,
            'alerts': monitor.get_alerts(),
            'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
    except Exception as e:
        logger.error(f"刷新基金数据失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/funds/add', methods=['POST'])
def add_fund():
    """添加监控基金"""
    try:
        data = request.get_json()
        fund_code = data.get('fund_code', '').strip()
        current_amount = float(data.get('current_amount', 0))
        max_amount = float(data.get('max_amount', 0))
        holding_return_rate = float(data.get('holding_return_rate', 0))

        if not fund_code:
            return jsonify({'success': False, 'message': '基金代码不能为空'}), 400

        success = monitor.add_fund(fund_code, current_amount, max_amount)

        if success:
            # 设置持有收益率
            if holding_return_rate != 0:
                monitor.update_holding_return_rate(fund_code, holding_return_rate)

            # 立即获取数据
            fund_data = monitor.get_fund_data(fund_code)
            return jsonify({
                'success': True,
                'message': f'添加成功: {fund_data.get("fund_name", fund_code)}',
                'data': fund_data
            })
        else:
            return jsonify({'success': False, 'message': '添加失败，请检查基金代码是否正确'}), 400

    except Exception as e:
        logger.error(f"添加基金失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/funds/<fund_code>', methods=['DELETE'])
def remove_fund(fund_code):
    """移除监控基金"""
    try:
        success = monitor.remove_fund(fund_code)
        if success:
            return jsonify({'success': True, 'message': '移除成功'})
        else:
            return jsonify({'success': False, 'message': '基金不存在'}), 404
    except Exception as e:
        logger.error(f"移除基金失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/funds/clear-all', methods=['DELETE'])
def clear_all_funds():
    """清除所有监控基金"""
    try:
        count = len(monitor.positions)
        monitor.clear_all_funds()
        return jsonify({
            'success': True,
            'message': f'已清除 {count} 只基金'
        })
    except Exception as e:
        logger.error(f"清除基金失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/funds/batch', methods=['POST'])
def batch_add_funds():
    """批量添加监控基金"""
    try:
        data = request.get_json()
        funds = data.get('funds', [])

        if not funds:
            return jsonify({'success': False, 'message': '没有基金数据'}), 400

        success_count = 0
        fail_count = 0
        results = []

        for fund in funds:
            fund_code = str(fund.get('fund_code', '')).strip()
            current_amount = float(fund.get('current_amount', 0) or 0)
            max_amount = float(fund.get('max_amount', 0) or 0)
            holding_return_rate = float(fund.get('holding_return_rate', 0) or 0)

            if not fund_code or not fund_code.isdigit() or len(fund_code) != 6:
                fail_count += 1
                results.append({
                    'fund_code': fund_code,
                    'success': False,
                    'message': '基金代码格式错误'
                })
                continue

            success = monitor.add_fund(fund_code, current_amount, max_amount)

            if success:
                if holding_return_rate != 0:
                    monitor.update_holding_return_rate(fund_code, holding_return_rate)
                success_count += 1
                results.append({
                    'fund_code': fund_code,
                    'success': True,
                    'message': '添加成功'
                })
            else:
                fail_count += 1
                results.append({
                    'fund_code': fund_code,
                    'success': False,
                    'message': '无效的基金代码'
                })

        return jsonify({
            'success': True,
            'message': f'成功添加 {success_count} 只基金，失败 {fail_count} 只',
            'data': {
                'success_count': success_count,
                'fail_count': fail_count,
                'results': results
            }
        })
    except Exception as e:
        logger.error(f"批量添加基金失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/funds/template', methods=['GET'])
def download_template():
    """下载批量添加基金模板"""
    try:
        import io
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "批量添加基金模板"

        # 表头
        headers = ['基金代码', '持仓金额', '上限金额', '持有收益率(%)']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # 示例数据
        examples = [
            ['000001', 10000, 50000, 5.5],
            ['110022', 5000, 20000, -2.3],
            ['519778', 8000, 30000, 10.2],
        ]
        for row_idx, example in enumerate(examples, 2):
            for col_idx, value in enumerate(example, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 说明
        ws.cell(row=5, column=1, value="【填写说明】")
        ws.cell(row=6, column=1, value="1. 基金代码：必填，6位数字")
        ws.cell(row=7, column=1, value="2. 持仓金额：选填，当前持仓金额")
        ws.cell(row=8, column=1, value="3. 上限金额：选填，持仓上限金额")
        ws.cell(row=9, column=1, value="4. 持有收益率：选填，当前持有收益率百分比")

        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='批量添加基金模板.xlsx'
        )
    except ImportError:
        return jsonify({'success': False, 'message': '请先安装 openpyxl: pip install openpyxl'}), 500
    except Exception as e:
        logger.error(f"下载模板失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/funds/upload', methods=['POST'])
def upload_funds_file():
    """上传Excel文件批量添加基金"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '没有上传文件'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '没有选择文件'}), 400

        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': '请上传Excel文件(.xlsx或.xls)'}), 400

        # 解析Excel
        from openpyxl import load_workbook
        wb = load_workbook(file)
        ws = wb.active

        funds = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # 跳过空行
                continue

            fund_code = str(row[0]).strip() if row[0] else ''
            current_amount = float(row[1]) if row[1] else 0
            max_amount = float(row[2]) if row[2] else 0
            holding_return_rate = float(row[3]) if row[3] else 0

            if fund_code:
                funds.append({
                    'fund_code': fund_code,
                    'current_amount': current_amount,
                    'max_amount': max_amount,
                    'holding_return_rate': holding_return_rate
                })

        if not funds:
            return jsonify({'success': False, 'message': 'Excel中没有有效的基金数据'}), 400

        # 批量添加
        success_count = 0
        fail_count = 0

        for fund in funds:
            fund_code = fund['fund_code']
            if monitor.add_fund(fund_code, fund['current_amount'], fund['max_amount']):
                if fund['holding_return_rate'] != 0:
                    monitor.update_holding_return_rate(fund_code, fund['holding_return_rate'])
                success_count += 1
            else:
                fail_count += 1

        return jsonify({
            'success': True,
            'message': f'成功添加 {success_count} 只基金，失败 {fail_count} 只',
            'data': {
                'success_count': success_count,
                'fail_count': fail_count
            }
        })
    except ImportError:
        return jsonify({'success': False, 'message': '请先安装 openpyxl: pip install openpyxl'}), 500
    except Exception as e:
        logger.error(f"上传文件失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/funds/<fund_code>/position', methods=['PUT'])
def update_position(fund_code):
    """更新持仓配置"""
    try:
        data = request.get_json()
        current_amount = data.get('current_amount')
        max_amount = data.get('max_amount')
        holding_return_rate = data.get('holding_return_rate')

        success = monitor.update_position(
            fund_code,
            current_amount=float(current_amount) if current_amount is not None else None,
            max_amount=float(max_amount) if max_amount is not None else None
        )

        # 更新持有收益率
        if holding_return_rate is not None and success:
            monitor.update_holding_return_rate(fund_code, float(holding_return_rate))

        if success:
            return jsonify({'success': True, 'message': '更新成功'})
        else:
            return jsonify({'success': False, 'message': '基金不存在'}), 404

    except Exception as e:
        logger.error(f"更新持仓失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/funds/<fund_code>/return_rate', methods=['PUT'])
def update_return_rate(fund_code):
    """更新持有收益率"""
    try:
        data = request.get_json()
        return_rate = data.get('holding_return_rate')

        if return_rate is None:
            return jsonify({'success': False, 'message': '持有收益率不能为空'}), 400

        success = monitor.update_holding_return_rate(fund_code, float(return_rate))

        if success:
            return jsonify({'success': True, 'message': '持有收益率已更新'})
        else:
            return jsonify({'success': False, 'message': '基金不存在'}), 404

    except Exception as e:
        logger.error(f"更新持有收益率失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/alerts', methods=['GET'])
def get_alerts():
    """获取提醒列表"""
    try:
        fund_code = request.args.get('fund_code')
        alerts = monitor.get_alerts(fund_code)
        return jsonify({
            'success': True,
            'data': alerts
        })
    except Exception as e:
        logger.error(f"获取提醒失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/alerts/clear', methods=['POST'])
def clear_alerts():
    """清除所有提醒"""
    try:
        monitor.clear_alerts()
        return jsonify({'success': True, 'message': '已清除所有提醒'})
    except Exception as e:
        logger.error(f"清除提醒失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/advices', methods=['GET'])
def get_advices():
    """获取所有基金的操作建议"""
    try:
        advisor_instance = get_advisor()
        advices = advisor_instance.generate_all_advices()
        return jsonify({
            'success': True,
            'data': [a.to_dict() for a in advices],
            'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
    except Exception as e:
        logger.error(f"获取建议失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/report', methods=['GET'])
def get_report():
    """获取下午报告"""
    try:
        advisor_instance = get_advisor()
        report = advisor_instance.generate_afternoon_report()
        return jsonify({
            'success': True,
            'data': report
        })
    except Exception as e:
        logger.error(f"获取报告失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ==================== 基金查询API ====================
# 注意：这些具体路由必须放在 /api/fund/<fund_code> 之前，否则会被通配路由拦截

@app.route('/api/fund/query/<fund_code>', methods=['GET'])
def query_fund(fund_code):
    """查询基金完整详情"""
    try:
        data = data_source_manager.get_fund_full_detail(fund_code)
        if not data:
            return jsonify({'success': False, 'message': '未找到该基金，请检查基金代码'}), 404

        return jsonify({
            'success': True,
            'data': data
        })
    except Exception as e:
        logger.error(f"查询基金失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/fund/history/<fund_code>', methods=['GET'])
def get_fund_history(fund_code):
    """获取基金历史净值数据"""
    try:
        period = request.args.get('period', '1y')
        data = data_source_manager.get_fund_history(fund_code, period)
        if not data:
            return jsonify({'success': False, 'message': '未找到历史数据'}), 404

        return jsonify({
            'success': True,
            'data': data
        })
    except Exception as e:
        logger.error(f"获取基金历史数据失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/fund/holdings/<fund_code>', methods=['GET'])
def get_fund_holdings(fund_code):
    """获取基金持仓数据"""
    try:
        data = data_source_manager.get_fund_holdings(fund_code)
        if not data:
            return jsonify({'success': False, 'message': '未找到持仓数据'}), 404

        return jsonify({
            'success': True,
            'data': data
        })
    except Exception as e:
        logger.error(f"获取基金持仓数据失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/fund/<fund_code>', methods=['GET'])
def get_fund_detail(fund_code):
    """获取单个基金详情"""
    try:
        data = monitor.get_fund_data(fund_code)
        if not data:
            return jsonify({'success': False, 'message': '基金不存在'}), 404

        position = monitor.get_position(fund_code)
        risk = monitor.get_risk_level(fund_code)

        result = {
            **data,
            'current_amount': position.current_amount if position else 0,
            'max_amount': position.max_amount if position else 0,
            'risk_level': risk
        }

        return jsonify({
            'success': True,
            'data': result
        })
    except Exception as e:
        logger.error(f"获取基金详情失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ==================== 数据源API ====================

@app.route('/api/datasource', methods=['GET'])
def get_datasource():
    """获取当前数据源和可用数据源列表"""
    try:
        return jsonify({
            'success': True,
            'data': {
                'current': data_source_manager.current_source,
                'current_name': data_source_manager.get_source_name(),
                'sources': data_source_manager.get_available_sources()
            }
        })
    except Exception as e:
        logger.error(f"获取数据源失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/datasource', methods=['PUT'])
def set_datasource():
    """切换数据源"""
    try:
        data = request.get_json()
        source_id = data.get('source_id', '').strip()

        if not source_id:
            return jsonify({'success': False, 'message': '数据源ID不能为空'}), 400

        if source_id not in data_source_manager.SOURCES:
            return jsonify({'success': False, 'message': f'不支持的数据源: {source_id}'}), 400

        old_source = data_source_manager.get_source_name()
        data_source_manager.current_source = source_id
        new_source = data_source_manager.get_source_name()

        # 清除缓存，下次获取数据时使用新数据源
        monitor.funds = {}

        logger.info(f"数据源已切换: {old_source} -> {new_source}")

        return jsonify({
            'success': True,
            'message': f'数据源已切换为: {new_source}',
            'data': {
                'current': source_id,
                'current_name': new_source
            }
        })
    except Exception as e:
        logger.error(f"切换数据源失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ==================== 定时任务 ====================

def scheduled_refresh():
    """定时刷新数据"""
    logger.info("执行定时刷新...")
    try:
        monitor.refresh_data()
    except Exception as e:
        logger.error(f"定时刷新失败: {e}")


def scheduled_afternoon_reminder():
    """下午提醒任务"""
    now = datetime.now()
    reminder_hour, reminder_minute = AFTERNOON_REMINDER_TIME

    if now.hour == reminder_hour and now.minute == reminder_minute:
        logger.info("执行下午提醒...")
        try:
            report = advisor.generate_afternoon_report()
            # 这里可以添加推送通知逻辑（如邮件、微信等）
            logger.info(f"下午报告: {report['summary']}")
        except Exception as e:
            logger.error(f"下午提醒失败: {e}")


# ==================== OCR识别API ====================

@app.route('/api/ocr/config', methods=['GET'])
def get_ocr_config():
    """获取OCR配置状态"""
    try:
        from ocr_service import ocr_service
        return jsonify({
            'success': True,
            'data': {
                'enabled': ocr_service.is_third_party_enabled(),
                'provider': ocr_service.get_provider_name()
            }
        })
    except Exception as e:
        logger.error(f"获取OCR配置失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/ocr/recognize', methods=['POST'])
def ocr_recognize():
    """OCR识别图片中的基金信息"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '没有上传图片'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '没有选择图片'}), 400

        # 检查文件类型
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp'}
        if not file.filename.lower().split('.')[-1] in allowed_extensions:
            return jsonify({'success': False, 'message': '不支持的图片格式'}), 400

        # 读取图片数据
        image_data = file.read()

        from ocr_service import ocr_service
        result = ocr_service.recognize_image(image_data)

        return jsonify(result.to_dict())
    except Exception as e:
        logger.error(f"OCR识别失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/ocr/parse', methods=['POST'])
def ocr_parse_text():
    """
    解析OCR识别的文本（前端Tesseract识别后调用）
    前端使用Tesseract.js识别后，将文本传给后端解析
    """
    try:
        data = request.get_json()
        text = data.get('text', '')

        if not text:
            return jsonify({'success': False, 'message': '没有识别文本'}), 400

        from ocr_service import FundParser
        parser = FundParser()
        funds = parser.parse_text(text)

        return jsonify({
            'success': True,
            'funds': funds,
            'message': f'识别完成，共发现 {len(funds)} 只基金'
        })
    except Exception as e:
        logger.error(f"解析OCR文本失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/fund/search', methods=['GET'])
def search_fund_by_name():
    """通过基金名称搜索基金代码"""
    try:
        keyword = request.args.get('keyword', '').strip()
        if not keyword or len(keyword) < 2:
            return jsonify({'success': False, 'message': '关键词太短'}), 400

        # 调用天天基金搜索接口
        import requests
        from urllib.parse import quote
        session = requests.Session()
        session.trust_env = False
        url = f"https://fundsuggest.eastmoney.com/FundSearch/api/FundSearchAPI.ashx?m=1&key={quote(keyword)}"
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Referer': 'https://fund.eastmoney.com/'
        }
        response = session.get(url, headers=headers, timeout=10)

        if response.status_code != 200:
            return jsonify({'success': False, 'message': '搜索失败'}), 500

        # 解析结果
        import json
        data = json.loads(response.text)
        datas = data.get('Datas', [])

        results = []
        for item in datas[:10]:  # 只返回前10条
            results.append({
                'fund_code': item.get('CODE', ''),
                'fund_name': item.get('NAME', ''),
                'fund_type': item.get('FundBaseInfo', {}).get('FTYPE', '')
            })

        return jsonify({
            'success': True,
            'data': results
        })
    except Exception as e:
        logger.error(f"搜索基金失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ==================== 启动应用 ====================

if __name__ == '__main__':
    print("=" * 50)
    print("基金实时监控助手 启动中...")
    print("=" * 50)
    print(f"刷新间隔: {REFRESH_INTERVAL}秒")
    print(f"下午提醒时间: {AFTERNOON_REMINDER_TIME[0]:02d}:{AFTERNOON_REMINDER_TIME[1]:02d}")
    print("=" * 50)
    print("访问地址: http://localhost:5000")
    print("=" * 50)

    # 启动Flask开发服务器
    app.run(
        host='0.0.0.0',
        port=5000,
        debug=True,
        threaded=True
    )
